import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from csv_to_excel import agent_csvs, csvs_to_excel


class CsvToExcelTests(unittest.TestCase):
    def write_csv(self, path, rows):
        with path.open("w", newline="", encoding="utf-8") as csv_file:
            writer = csv.DictWriter(
                csv_file,
                fieldnames=["Commercial", "AgencyName", "State"],
            )
            writer.writeheader()
            writer.writerows(rows)

    def test_commercial_output_filters_agents_csv_rows(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            vt_csv = temp_path / "vtAgents.csv"
            wy_csv = temp_path / "wyAgents.csv"
            old_commercial_csv = temp_path / "vtCommercialAgents.csv"

            self.write_csv(
                vt_csv,
                [
                    {"Commercial": "Y", "AgencyName": "A", "State": "VT"},
                    {"Commercial": "N", "AgencyName": "B", "State": "VT"},
                ],
            )
            self.write_csv(
                wy_csv,
                [{"Commercial": "Y", "AgencyName": "C", "State": "WY"}],
            )
            self.write_csv(
                old_commercial_csv,
                [{"Commercial": "Y", "AgencyName": "Ignored", "State": "VT"}],
            )

            selected_csvs = agent_csvs(str(temp_path))
            self.assertEqual(
                [Path(path).name for path in selected_csvs],
                ["vtAgents.csv", "wyAgents.csv"],
            )

            all_output = temp_path / "All.xlsx"
            commercial_output = temp_path / "Commercial.xlsx"
            csvs_to_excel(str(temp_path), str(all_output), selected_csvs)
            csvs_to_excel(
                str(temp_path),
                str(commercial_output),
                selected_csvs,
                commercial_only=True,
            )

            all_workbook = load_workbook(all_output, read_only=True)
            commercial_workbook = load_workbook(commercial_output, read_only=True)

            self.assertEqual(all_workbook.sheetnames, ["vt", "wy"])
            self.assertEqual(commercial_workbook.sheetnames, ["vt", "wy"])
            self.assertEqual(all_workbook["vt"].max_row - 1, 2)
            self.assertEqual(commercial_workbook["vt"].max_row - 1, 1)
            self.assertEqual(commercial_workbook["wy"].max_row - 1, 1)


if __name__ == "__main__":
    unittest.main()
